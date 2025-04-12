import streamlit as st
import pandas as pd
from data_loader import load_data
from jadwal_model import Jadwal
from genetic_algorithm import AlgoritmaGenetika
from tabu_search import TabuSearch
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import json
import logging
from datetime import datetime

class AppLogger:
    def __init__(self):
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)
        
    def log(self, message):
        self.logger.info(message)
        st.info(message)

logger = AppLogger()

def display_schedule(jadwal):
    """
    Menampilkan jadwal dalam bentuk tabel di Streamlit.
    """
    if not jadwal:
        st.warning("Jadwal kosong")
        return
        
    df = pd.DataFrame([{
        "Hari": j.hari,
        "Mulai": j.waktu_mulai,
        "Selesai": j.waktu_selesai,
        "Mata Kuliah": j.matakuliah["nama"],
        "Dosen": j.dosen,
        "Ruang": j.ruang,
        "Kelas": j.kelas,
        "Semester": j.semester
    } for j in jadwal])
    
    # Tampilkan statistik jadwal
    st.subheader("Statistik Jadwal")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Total Mata Kuliah", len(jadwal))
    
    with col2:
        unique_rooms = len(set(j.ruang for j in jadwal))
        st.metric("Ruang yang Digunakan", unique_rooms)
    
    with col3:
        unique_teachers = len(set(j.dosen for j in jadwal))
        st.metric("Dosen yang Mengajar", unique_teachers)
    
    # Tampilkan jadwal per semester dan hari
    st.subheader("Jadwal Kuliah")
    
    for semester in sorted(df["Semester"].unique()):
        semester_data = df[df["Semester"] == semester]
        
        with st.expander(f"Semester {semester}"):
            for hari in ["Senin", "Selasa", "Rabu", "Kamis", "Jumat"]:
                hari_data = semester_data[semester_data["Hari"] == hari]
                if not hari_data.empty:
                    st.subheader(f"{hari}")
                    st.dataframe(hari_data)

def export_to_excel(jadwal):
    """
    Mengekspor jadwal ke file Excel dengan formatting.
    """
    if not jadwal:
        return None
        
    df = pd.DataFrame([{
        "Hari": j.hari,
        "Mulai": j.waktu_mulai,
        "Selesai": j.waktu_selesai,
        "Mata Kuliah": j.matakuliah["nama"],
        "Dosen": j.dosen,
        "Ruang": j.ruang,
        "Kelas": j.kelas,
        "Semester": j.semester
    } for j in jadwal])
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for semester in sorted(df['Semester'].unique()):
            semester_data = df[df['Semester'] == semester]
            semester_data_sorted = []
            
            for hari in ["Senin", "Selasa", "Rabu", "Kamis", "Jumat"]:
                hari_data = semester_data[semester_data['Hari'] == hari]
                if not hari_data.empty:
                    hari_data['Hari'] = hari  # Tambahkan label hari
                    semester_data_sorted.append(hari_data)
                    semester_data_sorted.append(pd.DataFrame())  # Tambahkan baris kosong untuk pemisah
            
            if semester_data_sorted:
                semester_data_sorted = pd.concat(semester_data_sorted)
                semester_data_sorted.to_excel(writer, sheet_name=f'Semester {semester}', index=False)
                
                # Formatting Excel
                workbook = writer.book
                worksheet = writer.sheets[f'Semester {semester}']
                
                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                              top=Side(style='thin'), bottom=Side(style='thin'))
                
                # Format header
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                
                # Format sel
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
    
    processed_data = output.getvalue()
    return processed_data

def main():
    """
    Fungsi utama untuk menjalankan aplikasi.
    """
    st.title("Aplikasi Penjadwalan Mata Kuliah")
    
    # Sidebar untuk konfigurasi
    st.sidebar.header("Konfigurasi")
    
    # Konfigurasi Genetic Algorithm
    with st.sidebar.expander("Genetic Algorithm"):
        ga_population = st.number_input("Ukuran Populasi", min_value=10, max_value=100, value=50)
        ga_generations = st.number_input("Jumlah Generasi", min_value=10, max_value=500, value=100)
        ga_crossover = st.slider("Probabilitas Crossover", 0.0, 1.0, 0.8)
        ga_mutation = st.slider("Probabilitas Mutasi", 0.0, 1.0, 0.1)
        ga_elitism = st.slider("Persentase Elitisme", 0.0, 1.0, 0.1)
    
    # Konfigurasi Tabu Search
    with st.sidebar.expander("Tabu Search"):
        ts_iterations = st.number_input("Iterasi Maksimum", min_value=10, max_value=500, value=100)
        ts_tabu_size = st.number_input("Ukuran Daftar Tabu", min_value=5, max_value=50, value=10)
        ts_aspiration = st.slider("Threshold Aspirasi", 0.0, 1.0, 0.1)
        ts_diversification = st.slider("Threshold Diversifikasi", 0.0, 1.0, 0.8)
        ts_intensification = st.slider("Threshold Intensifikasi", 0.0, 1.0, 0.9)
    
    # Memuat data
    try:
        data = load_data()
        logger.log("Data berhasil dimuat")
    except Exception as e:
        logger.log(f"Error saat memuat data: {str(e)}")
        st.error(f"Error: {str(e)}")
        return
    
    # Tab untuk algoritma
    tab1, tab2 = st.tabs(["Genetic Algorithm", "Tabu Search"])
    
    with tab1:
        if st.button("Optimalkan Jadwal (Genetic Algorithm)"):
            with st.spinner("Sedang menjalankan algoritma genetika..."):
                try:
                    start_time = datetime.now()
                    
                    # Inisialisasi dan jalankan algoritma
                    ag = AlgoritmaGenetika(
                        data,
                        ukuran_populasi=ga_population,
                        generasi=ga_generations,
                        crossover_rate=ga_crossover,
                        mutation_rate=ga_mutation,
                        elitism_rate=ga_elitism
                    )
                    
                    best_solution, best_fitness, metrics = ag.run()
                    
                    end_time = datetime.now()
                    execution_time = (end_time - start_time).total_seconds()
                    
                    # Tampilkan hasil
                    st.success(f"Optimasi selesai dalam {execution_time:.2f} detik!")
                    st.write(f"Fitness Terbaik: {best_fitness:.4f}")
                    
                    # Tampilkan statistik
                    st.subheader("Statistik Optimasi")
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Generasi", ga_generations)
                    
                    with col2:
                        st.metric("Populasi", ga_population)
                    
                    with col3:
                        st.metric("Fitness", f"{best_fitness:.4f}")
                    
                    # Tampilkan jadwal
                    display_schedule(best_solution)
                    
                    # Tambahkan tombol download
                    excel_data = export_to_excel(best_solution)
                    if excel_data:
                        st.download_button(
                            label="Unduh Jadwal sebagai Excel",
                            data=excel_data,
                            file_name=f'jadwal_genetic_algorithm_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                except Exception as e:
                    logger.log(f"Error saat menjalankan Genetic Algorithm: {str(e)}")
                    st.error(f"Error: {str(e)}")
    
    with tab2:
        if st.button("Optimalkan Jadwal (Tabu Search)"):
            with st.spinner("Sedang menjalankan Tabu Search..."):
                try:
                    start_time = datetime.now()
                    
                    # Inisialisasi algoritma
                    initial_solution = AlgoritmaGenetika(data).generate_random_jadwal()
                    ts = TabuSearch(
                        lambda x: 1 / (1 + sum(1 for i in range(len(x)) for j in range(i + 1, len(x)) 
                                   if isinstance(x[i], Jadwal) and isinstance(x[j], Jadwal) 
                                   and x[i].hari == x[j].hari and x[i].ruang == x[j].ruang 
                                   and ((x[i].waktu_mulai <= x[j].waktu_mulai < x[i].waktu_selesai) 
                                        or (x[j].waktu_mulai <= x[i].waktu_mulai < x[j].waktu_selesai)))),
                        initial_solution,
                        max_iter=ts_iterations,
                        tabu_size=ts_tabu_size,
                        aspiration_threshold=ts_aspiration,
                        diversification_threshold=ts_diversification,
                        intensification_threshold=ts_intensification
                    )
                    
                    best_solution, best_fitness, metrics = ts.search()
                    
                    end_time = datetime.now()
                    execution_time = (end_time - start_time).total_seconds()
                    
                    # Tampilkan hasil
                    st.success(f"Optimasi selesai dalam {execution_time:.2f} detik!")
                    st.write(f"Fitness Terbaik: {best_fitness:.4f}")
                    
                    # Tampilkan statistik
                    st.subheader("Statistik Optimasi")
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Iterasi", ts_iterations)
                    
                    with col2:
                        st.metric("Ukuran Tabu", ts_tabu_size)
                    
                    with col3:
                        st.metric("Fitness", f"{best_fitness:.4f}")
                    
                    # Tampilkan jadwal
                    display_schedule(best_solution)
                    
                    # Tambahkan tombol download
                    excel_data = export_to_excel(best_solution)
                    if excel_data:
                        st.download_button(
                            label="Unduh Jadwal sebagai Excel",
                            data=excel_data,
                            file_name=f'jadwal_tabu_search_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                except Exception as e:
                    logger.log(f"Error saat menjalankan Tabu Search: {str(e)}")
                    st.error(f"Error: {str(e)}")

if __name__ == "__main__":
    main()
